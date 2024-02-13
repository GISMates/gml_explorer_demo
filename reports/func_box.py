import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side

from qgis.core import QgsProject, NULL
import qgis

from qgis.PyQt.QtCore import Qt, QSize
from qgis.PyQt.QtWidgets import QDockWidget, QWidget, QTableWidget, QHeaderView, QTableWidgetItem, QPushButton, \
    QFileDialog, QMessageBox, QVBoxLayout, QHBoxLayout, QSpacerItem, QSizePolicy, QTextBrowser, QMenu

iface = qgis.utils.iface
project = QgsProject().instance()


def configureDialog(parent, title, columns, layer):
    dialog = QDockWidget(parent)
    dialog.setWindowTitle(title)
    table = QTableWidget(0, len(columns))
    container = QWidget()
    layout = QVBoxLayout()
    button_layout = QHBoxLayout()
    csv_button = createButton(dialog, 'Eksport do csv', lambda: toCsv(dialog, table, title, columns))
    excel_button = createButton(dialog, 'Eksport do excela', lambda: toExcel(dialog, table, title, columns))
    button_layout.addWidget(csv_button)
    button_layout.addWidget(excel_button)
    button_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
    layout.addLayout(button_layout)
    horizontal_header = table.horizontalHeader()
    horizontal_header.setSectionResizeMode(QHeaderView.Stretch)
    horizontal_header.hideSection(len(columns) - 1)
    table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
    table.setHorizontalHeaderLabels(columns)
    table.itemSelectionChanged.connect(lambda: zoomToSelected(layer, table))
    layout.addWidget(table)
    container.setLayout(layout)
    dialog.setWidget(container)
    return dialog, table


def configureLandReportDialog(parent, title, columns, layer):
    dialog = QDockWidget(parent)
    dialog.setWindowTitle(title)
    container = QWidget()
    layout = QHBoxLayout()
    table = QTableWidget(0, len(columns))
    table_layout = QVBoxLayout()
    button_layout = QHBoxLayout()
    export_button = createButton(dialog, 'Eksport do html', lambda: toHtml(dialog, table, title))
    excel_button = createButton(dialog, 'Eksport do excela', lambda: toExcel(dialog, table, title, columns, True))
    button_layout.addWidget(export_button)
    button_layout.addWidget(excel_button)
    button_layout.addItem(QSpacerItem(40, 20, QSizePolicy.Expanding, QSizePolicy.Minimum))
    report_browser = QTextBrowser()
    report_browser.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
    table_layout.addLayout(button_layout)
    horizontal_header = table.horizontalHeader()
    horizontal_header.setSectionResizeMode(QHeaderView.Stretch)
    table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
    table.setHorizontalHeaderLabels(columns)
    table.itemSelectionChanged.connect(lambda: zoomToSelected(layer, table))
    table.itemSelectionChanged.connect(lambda: displayReport(table, report_browser))
    table_layout.addWidget(table)
    layout.addLayout(table_layout)
    layout.addWidget(report_browser)
    container.setLayout(layout)
    dialog.setWidget(container)
    report_browser.contextMenuEvent = lambda event: browserMenu(event, report_browser, dialog, table)
    dialog.resizeEvent = lambda event: adjustSize(event, report_browser)
    return dialog, table


def browserMenu(event, report_browser, dialog, table):
    if report_browser.toPlainText() != '':
        menu = QMenu()
        menu.addAction('Zapisz dokument')
        menu.addAction('Zapisz dokument (excel)')
        action = menu.exec(event.globalPos())
        if action:
            if action.text() == 'Zapisz dokument':
                toHtml(dialog, table, 'Informacja z rejestru gruntów', True)
            if action.text() == 'Zapisz dokument (excel)':
                toExcel(dialog, table, 'Informacja z rejestru gruntów', [], True, True)


def adjustSize(event, report_browser):
    width = int(event.size().width() * 0.75)
    report_browser.setMinimumWidth(width)
    report_browser.setMaximumWidth(width)


def createButton(dialog, text, func):
    button = QPushButton(dialog)
    button.setText(text)
    button.clicked.connect(func)
    return button


def zoomToSelected(layer, table):
    layer.removeSelection()
    layer.select(int(table.item(table.currentRow(), 0).whatsThis()))
    iface.mapCanvas().zoomToSelected(layer)


def displayReport(table, report_browser):
    report_browser.setHtml(table.currentItem().data(100))


def showMessage(dialog, title, text):
    mbox = QMessageBox(dialog)
    mbox.setWindowTitle(title)
    mbox.setText(text)
    mbox.exec()


def toHtml(dialog, table, title, current=False):
    html_filepath = QFileDialog.getSaveFileName(dialog, 'Wskaż ścieżkę zapisu', '', 'HyperText Markup Language (*.html)')[0]
    if html_filepath != '':
        with open(html_filepath, 'w') as f:
            if current:
                f.write(table.currentItem().data(100))
            else:
                f.write(table.item(0, 0).data(101))
        showMessage(dialog, title,
                    '<p>Dokument został zapisany w lokalizacji:</p>'
                    '<p><i>%s</i></p>' % html_filepath)


def toCsv(dialog, table, title, columns):
    path, ok = QFileDialog.getSaveFileName(dialog, 'Wybierz plik', '', 'Comma Separated Values (*.csv)')
    if ok:
        df = prepareExportData(table, columns)
        df.to_csv(path, sep=';')
        showMessage(dialog, title,
                    '<p>Zestawienie zostało zapisane w lokalizacji:</p>'
                    '<p><i>%s</i></p>' % path)


def toExcel(dialog, table, title, columns, jrg=False, current=False):
    path, ok = QFileDialog.getSaveFileName(dialog, 'Wybierz plik', '', 'Excel (*.xlsx)')
    if ok:
        if jrg:
            wb = Workbook()
            ws = wb.active
            r = 1
            if current:
                r = itemExcel(ws, table.currentItem(), r)
            else:
                for i in range(0, table.rowCount()):
                    item = table.item(i, 0)
                    r = itemExcel(ws, item, r)
                    r += 2
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                for cell in column_cells:
                    if cell.value != 'INFORMACJA Z REJESTRU GRUNTÓW':
                        cell.alignment = Alignment(wrapText=True, vertical='center')
                    bold = cell.font.bold
                    size = cell.font.size
                    italic = cell.font.italic
                    if bold and size == 12:
                        size = 12
                    else:
                        size = 10
                    cell.font = Font(size=size, bold=bold, italic=italic)
                length = 30 if length > 30 else length
                column = get_column_letter(column_cells[0].column)
                ws.column_dimensions[column].width = length
            wb.save(path)
        else:
            df = prepareExportData(table, columns)
            with pd.ExcelWriter(path) as writer:
                df.to_excel(writer, sheet_name=title, index=False)
        showMessage(dialog, title,
                    '<p>Zestawienie zostało zapisane w lokalizacji:</p>'
                    '<p><i>%s</i></p>' % path)


def itemExcel(ws, item, r):
    report_data = item.data(200)
    field_info = report_data[0]
    location_dict = report_data[1]
    landuses_info = report_data[2]
    ownership_info = report_data[3]
    docs_info = report_data[4]
    buildings_info = report_data[5]
    dicts = report_data[6]
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws.title = 'Informacja z rejestru gruntów'
    ws[f'E{r}'] = 'INFORMACJA Z REJESTRU GRUNTÓW'
    ws[f'E{r}'].font = Font(bold=True, size=12)
    for info in field_info:
        r += 1
        ws[f'I{r}'] = info
        ws[f'I{r}'].font = Font(bold=True, size=12)
    r += 2
    headers_dict = {'A': ['Jedn. ew.', 'Obręb', 'Ident.', 'Pow. ew.', 'Woj.', 'Powiat', 'Gmina', 'Nr KW'],
                    'D': ['Jedn. rej.', 'Grupa rej.', 'Adres', '', 'Wydruk z dn.', 'Uwagi']}
    for letter, columns in headers_dict.items():
        hr = r
        for col in columns:
            ws[f'{letter}{hr}'] = col
            ws[f'{letter}{hr}'].font = Font(italic=True)
            hr += 1
    for letter, columns in location_dict.items():
        hr = r
        for col in columns:
            ws[f'{letter}{hr}'] = col
            hr += 1
    r = hr + 3
    if len(ownership_info) > 0:
        header = ['Właściciel', 'Adres', 'Rodzaj prawa', 'Udział']
        for i, h in enumerate(header, 1):
            ws[f'{get_column_letter(i)}{r}'] = h
            ws[f'{get_column_letter(i)}{r}'].font = Font(bold=True)
            ws[f'{get_column_letter(i)}{r}'].border = thin_border
        for info in ownership_info:
            r += 1
            for i in range(0, len(info)):
                ws[f'{get_column_letter(i + 1)}{r}'] = info[i].replace('<br>', '\n')
                ws[f'{get_column_letter(i + 1)}{r}'].border = thin_border
    if len(landuses_info) > 0:
        r += 2
        ws[f'A{r}'] = 'KLASOUŻYTKI'
        ws[f'A{r}'].font = Font(bold=True, size=12)
        r += 1
        header = ['Sposób zagospod.', 'Rodzaj użytku', 'Klasa bonitacyjna', 'Powierzchnia ewidencyjna']
        for i, h in enumerate(header, 1):
            ws[f'{get_column_letter(i)}{r}'] = h
            ws[f'{get_column_letter(i)}{r}'].font = Font(bold=True)
            ws[f'{get_column_letter(i)}{r}'].border = thin_border
        landuses_area = 0
        for info in landuses_info:
            r += 1
            for i in range(0, len(info)):
                ws[f'{get_column_letter(i + 1)}{r}'] = info[i]
                ws[f'{get_column_letter(i + 1)}{r}'].border = thin_border
            try:
                landuse_area = round(float(info[3]), 4)
            except:
                landuse_area = 0
            landuses_area += landuse_area
        r += 1
        ws[f'C{r}'] = 'Suma powierzchni'
        ws[f'C{r}'].border = thin_border
        ws[f'D{r}'] = str(landuses_area)
        ws[f'D{r}'].border = thin_border
    if len(docs_info) > 0:
        r += 2
        ws[f'A{r}'] = 'DOKUMENTY'
        ws[f'A{r}'].font = Font(bold=True, size=12)
        r += 1
        header = ['Typ', 'Rodzaj', 'Data dok./przek. do zas.', 'Sygnatura/ozn. kanc.', 'Nazwa twórcy', 'Opis dokumentu']
        for i, h in enumerate(header, 1):
            ws[f'{get_column_letter(i)}{r}'] = h
            ws[f'{get_column_letter(i)}{r}'].font = Font(bold=True)
            ws[f'{get_column_letter(i)}{r}'].border = thin_border
        for info in docs_info:
            r += 1
            for i in range(0, len(info)):
                ws[f'{get_column_letter(i + 1)}{r}'] = info[i]
                ws[f'{get_column_letter(i + 1)}{r}'].border = thin_border
    if len(buildings_info) > 0:
        r += 2
        ws[f'A{r}'] = 'BUDYNKI'
        ws[f'A{r}'].font = Font(bold=True, size=12)
        r += 1
        header = ['Ident.', 'Status', 'PKOB', 'FSB\nKST', 'Nr KW', 'Mat. ścian', 'Kond. nadz.\nKond. podz.',
                  'P. zab. (m2).', 'P. uż. z obm. (m2).', 'Rok zak. bud.', 'Adres budynku\nNr rej. zabytków']
        for i, h in enumerate(header, 1):
            ws[f'{get_column_letter(i)}{r}'] = h
            ws[f'{get_column_letter(i)}{r}'].font = Font(bold=True)
            ws[f'{get_column_letter(i)}{r}'].border = thin_border
        for info in buildings_info:
            r += 1
            info = list(info)
            info[3] = f'{info[3]}\n{info[4]}'
            info[7] = f'{info[7]}\n{info[8]}'
            info.remove(info[4])
            info.remove(info[7])
            for i in range(0, len(info)):
                ws[f'{get_column_letter(i + 1)}{r}'] = info[i]
                ws[f'{get_column_letter(i + 1)}{r}'].border = thin_border
        if len(dicts) > 0:
            r += 2
            ws[f'A{r}'] = 'Oznaczenia'
            ws[f'A{r}'].font = Font(bold=True, size=12)
            for attr, attr_mapping in dicts.items():
                r += 1
                ws[f'A{r}'] = attr
                for k, v in attr_mapping.items():
                    r += 1
                    ws[f'A{r}'] = k
                    ws[f'B{r}'] = v
    return r


def prepareExportData(table, columns):
    data = []
    for row in range(table.rowCount()):
        rowData = []
        for column in range(table.columnCount()):
            widgetItem = table.item(row, column)
            if widgetItem and widgetItem.text:
                rowData.append(widgetItem.text())
            else:
                rowData.append('NULL')
        data.append(rowData)
    df = pd.DataFrame(data)
    df.columns = columns
    return df


def getLayer(lname, message):
    try:
        return project.mapLayersByName(lname)[0], message
    except:
        message += '<p>Do projektu nie została wczytana warstwa %s</p>' % lname
        return None, message


def getRelation(relations_layer, lname, relname, one_to_many=False):
    if one_to_many:
        relation = [None, None]
    else:
        relation = ''
    for relation_feature in relations_layer.getFeatures("Warstwa = '%s' AND Relacja = '%s'" % (lname, relname)):
        attribute = relation_feature.attribute("Atrybut")
        if one_to_many:
            if relation[0]:
                relation[1] = attribute
            else:
                relation[0] = attribute
        else:
            relation = attribute
    return relation


def getFeatureValue(f, a):
    try:
        v = f.attribute(a)
        if v == NULL or v.strip() == '':
            v = ' '
    except:
        v = ' '
    return v


def getListValue(l, i):
    try:
        return l[i]
    except:
        return ' '


def fillDict(dicts_container, f, v, m):
    d = dicts_container.get(f, {})
    if m.get(v):
        d[v] = m.get(v)
        dicts_container[f] = d
